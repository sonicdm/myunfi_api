from pathlib import Path
import csv
from openpyxl import load_workbook, Workbook
import logging

logger = logging.getLogger(__name__)

def read_csv_to_dict(file_path):
    """
    Reads a csv file to a dictionary.
    """
    logger.debug('Reading csv file: {}'.format(file_path))
    with open(file_path, 'r') as csv_file:
        reader = csv.DictReader(csv_file)
        return [row for row in reader]

def read_csv_to_list(file_path):
    """
    Reads a csv file to a list.
    """
    logger.debug('Reading csv file: {}'.format(file_path))
    with open(file_path, 'r') as csv_file:
        reader = csv.reader(csv_file)
        return [row for row in reader]
    
def read_csv(file_path, dict_reader=True):
    """
    Reads a csv file to a dictionary or list.
    """
    if dict_reader:
        return read_csv_to_dict(file_path)
    else:
        return read_csv_to_list(file_path)
    

def read_xlsx(file_path):
    """
    Reads an excel file to a list of lists.
    """
    wb = load_workbook(file_path)
    ws = wb.active
    return [[cell.value for cell in row] for row in ws.rows]

def load_xlsx(file_path):
    """
    Loads an excel file to a list of lists.
    """
    wb = load_workbook(file_path)
    return wb